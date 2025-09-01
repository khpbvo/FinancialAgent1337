from __future__ import annotations

import hashlib
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable


@dataclass
class IngestStats:
    docs_new: int = 0
    tx_seen: int = 0
    tx_new: int = 0


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _rows_from_xls(path: Path) -> tuple[list[str], Iterable[list[Any]]]:
    try:
        import xlrd  # type: ignore

        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_index(0)
        headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
        def gen():
            for r in range(1, sheet.nrows):
                yield [sheet.cell_value(r, c) for c in range(sheet.ncols)]
        return headers, gen()
    except ModuleNotFoundError:
        raise


def _rows_from_xlsx(path: Path) -> tuple[list[str], Iterable[list[Any]]]:
    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h is not None else "" for h in (rows[0] if rows else [])]
        return headers, (list(r) for r in rows[1:])
    except ModuleNotFoundError:
        raise


def ingest_xls(conn: sqlite3.Connection, xls_path: Path) -> IngestStats:
    from .csv_ingest import ingest_csv_ing  # reuse logic by writing temp CSV if needed
    from .normalization import (
        parse_eu_amount_to_cents,
        try_extract_value_date,
        compute_txn_hash,
        TxnHashInput,
    )
    from .csv_ingest import _ensure_account  # type: ignore[attr-defined]

    stats = IngestStats()
    sha = _sha256_file(xls_path)
    cur = conn.execute("SELECT id FROM documents WHERE sha256=?", (sha,))
    row = cur.fetchone()
    if row:
        doc_id = int(row[0])
    else:
        cur = conn.execute(
            "INSERT INTO documents(path, sha256, source_type, bank_hint) VALUES (?,?,?,?)",
            (str(xls_path), sha, xls_path.suffix.lower().lstrip("."), "ING"),
        )
        stats.docs_new += 1
        doc_id = int(cur.lastrowid)

    # Load rows
    headers: list[str]
    rows: Iterable[list[Any]]
    try:
        if xls_path.suffix.lower() == ".xlsx":
            headers, rows = _rows_from_xlsx(xls_path)
        else:
            headers, rows = _rows_from_xls(xls_path)
    except ModuleNotFoundError as e:
        conn.execute(
            "INSERT INTO parse_events(document_id, stage, ok, message) VALUES (?,?,?,?)",
            (doc_id, "parse", 0, f"Excel parser missing: {e}"),
        )
        conn.commit()
        return stats

    # Standardize header names (strip/casefold)
    norm_headers = [h.strip().casefold() for h in headers]

    def idx(name_variants: list[str]) -> int | None:
        for v in name_variants:
            if v in norm_headers:
                return norm_headers.index(v)
        return None

    idx_datum = idx(["datum", "date"])
    idx_name = idx(["naam / omschrijving", "naam/omschrijving", "omschrijving", "description"]) 
    idx_rek = idx(["rekening", "iban", "account"])
    idx_tegen = idx(["tegenrekening", "counterparty", "iban tegenrekening"]) 
    idx_afbij = idx(["af bij", "af/bij", "sign"]) 
    idx_bedrag = idx(["bedrag (eur)", "bedrag", "amount", "amount (eur)"]) 
    idx_mut = idx(["mutatiesoort", "type"]) 
    idx_meded = idx(["mededelingen", "details", "memo"]) 

    for r in rows:
        stats.tx_seen += 1
        try:
            def cell(i: int | None) -> str:
                if i is None or i >= len(r) or i < 0:
                    return ""
                v = r[i]
                return str(v).strip() if v is not None else ""

            raw_date = cell(idx_datum)
            if raw_date and raw_date.isdigit() and len(raw_date) == 8:
                booking_iso = f"{raw_date[0:4]}-{raw_date[4:6]}-{raw_date[6:8]}"
            else:
                # Attempt Excel serial date
                try:
                    import xlrd  # type: ignore

                    if isinstance(r[idx_datum or 0], (int, float)):
                        dtuple = xlrd.xldate_as_tuple(r[idx_datum or 0], 0)
                        booking_iso = f"{dtuple[0]:04d}-{dtuple[1]:02d}-{dtuple[2]:02d}"
                    else:
                        booking_iso = raw_date
                except Exception:
                    booking_iso = raw_date

            description = cell(idx_meded) or cell(idx_name)
            value_date = try_extract_value_date(description) or None
            debit_credit = cell(idx_afbij) or "Af"  # default to debit if unknown
            amount_cents = parse_eu_amount_to_cents(cell(idx_bedrag) or "0", debit_credit)
            currency = "EUR"
            counterparty_iban = cell(idx_tegen) or None
            counterparty_name = cell(idx_name) or None
            account_str = cell(idx_rek) or ""
            account_id = _ensure_account(conn, iban=account_str or None, account_no=None)

            from .normalization import normalize_description as norm_desc_fn
            norm_desc = norm_desc_fn(description)
            cp_ref = counterparty_iban or (counterparty_name or "")
            from .normalization import compute_txn_hash as compute_hash_fn, TxnHashInput as TxnHashInputCls
            txn_hash = compute_hash_fn(
                TxnHashInputCls(
                    account_id=account_id,
                    booking_date=booking_iso,
                    value_date=value_date,
                    amount_cents=amount_cents,
                    currency=currency,
                    debit_credit=("DEBIT" if debit_credit.lower() == "af" else "CREDIT"),
                    counterparty_iban_or_name=cp_ref,
                    normalized_description=norm_desc,
                )
            )
            conn.execute(
                """
                INSERT OR IGNORE INTO transactions(
                    account_id, document_id, txn_hash, booking_date, value_date,
                    amount_cents, currency, debit_credit, counterparty_name,
                    counterparty_iban, description
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    account_id,
                    doc_id,
                    txn_hash,
                    booking_iso,
                    value_date,
                    amount_cents,
                    currency,
                    ("DEBIT" if debit_credit.lower() == "af" else "CREDIT"),
                    counterparty_name,
                    counterparty_iban,
                    description,
                ),
            )
            ch = conn.execute("SELECT changes()").fetchone()[0]
            if int(ch) > 0:
                stats.tx_new += 1
        except Exception as e:  # noqa: BLE001
            conn.execute(
                "INSERT INTO parse_events(document_id, stage, ok, message) VALUES (?,?,?,?)",
                (doc_id, "parse", 0, f"xls row error: {type(e).__name__}: {e}"),
            )

    conn.execute(
        "INSERT INTO parse_events(document_id, stage, ok, message) VALUES (?,?,?,?)",
        (doc_id, "parse", 1, f"excel rows: seen={stats.tx_seen}, new={stats.tx_new}"),
    )
    conn.commit()
    return stats
